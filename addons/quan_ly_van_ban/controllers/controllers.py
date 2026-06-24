# -*- coding: utf-8 -*-
from odoo import http
from odoo.http import request

import io
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


class TimKiemController(http.Controller):

    @http.route('/qlvb/search', type='json', auth='user', methods=['POST'])
    def tim_kiem(self, tu_khoa='', **kw):
        if not tu_khoa or len(tu_khoa.strip()) < 2:
            return {'results': [], 'message': 'Vui lòng nhập ít nhất 2 ký tự.'}

        tu_khoa = tu_khoa.strip()
        results = []

        # 1. Văn bản đi – dùng '|' để OR các điều kiện
        van_ban_di = request.env['van_ban_di'].search([
            '|', '|',
            ('tieu_de', 'ilike', tu_khoa),
            ('trich_yeu', 'ilike', tu_khoa),
            ('so_hieu', 'ilike', tu_khoa),
        ], limit=5)
        for vb in van_ban_di:
            results.append({
                'loai': '📄 Văn bản đi',
                'ten': vb.tieu_de,
                'mo_ta': f"Số hiệu: {vb.so_hieu or ''} | Ngày: {vb.ngay_di or ''}",
                'id': vb.id,
                'model': 'van_ban_di',
                'action': 'action_van_ban_di',
            })

        # 2. Văn bản đến
        van_ban_den = request.env['van_ban_den'].search([
            '|',
            ('tieu_de', 'ilike', tu_khoa),
            ('trich_yeu', 'ilike', tu_khoa),
        ], limit=5)
        for vb in van_ban_den:
            results.append({
                'loai': '📨 Văn bản đến',
                'ten': vb.tieu_de,
                'mo_ta': f"Ngày đến: {vb.ngay_den or ''}",
                'id': vb.id,
                'model': 'van_ban_den',
                'action': 'action_van_ban_den',
            })

        # 3. Công việc
        cong_viec = request.env['cong.viec'].search([
            '|',
            ('ten_cong_viec', 'ilike', tu_khoa),
            ('yeu_cau', 'ilike', tu_khoa),
        ], limit=5)
        for cv in cong_viec:
            results.append({
                'loai': '✅ Công việc',
                'ten': cv.ten_cong_viec,
                'mo_ta': f"Hạn: {cv.han_xu_ly or ''} | Chủ trì: {cv.chu_tri_giai_quyet.ho_ten if cv.chu_tri_giai_quyet else ''}",
                'id': cv.id,
                'model': 'cong.viec',
                'action': 'action_cong_viec',
            })

        # 4. Khách hàng
        khach_hang = request.env['khach_hang'].search([
            '|', '|',
            ('ten_khach_hang', 'ilike', tu_khoa),
            ('ma_khach_hang', 'ilike', tu_khoa),
            ('so_dien_thoai', 'ilike', tu_khoa),
        ], limit=5)
        for kh in khach_hang:
            results.append({
                'loai': '👤 Khách hàng',
                'ten': kh.ten_khach_hang,
                'mo_ta': f"Mã: {kh.ma_khach_hang or ''} | SĐT: {kh.so_dien_thoai or ''}",
                'id': kh.id,
                'model': 'khach_hang',
                'action': 'action_khach_hang',
            })

        message = f"Tìm thấy {len(results)} kết quả." if results else "Không tìm thấy kết quả nào."
        return {'results': results, 'message': message}


class BaoCaoController(http.Controller):

    @http.route('/qlvb/export/excel', type='http', auth='user', methods=['GET'])
    def export_excel(self, thang=None, nam=None, loai='all', **kw):
        """Xuất báo cáo Excel thống kê văn bản và công việc theo tháng"""

        now = datetime.now()
        try:
            thang = int(thang) if thang else now.month
            nam = int(nam) if nam else now.year
        except ValueError:
            thang, nam = now.month, now.year

        wb = openpyxl.Workbook()

        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill_xh = PatternFill('solid', fgColor='1F4E79')
        header_fill_cv = PatternFill('solid', fgColor='375623')
        center = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left = Alignment(horizontal='left', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )
        title_font = Font(bold=True, size=13, color='1F4E79')

        def style_header_row(ws, row, col_count, fill):
            for c in range(1, col_count + 1):
                cell = ws.cell(row=row, column=c)
                cell.font = header_font
                cell.fill = fill
                cell.alignment = center
                cell.border = thin_border

        def style_data_row(ws, row, col_count, alt=False):
            alt_fill = PatternFill('solid', fgColor='DEEAF1') if alt else None
            for c in range(1, col_count + 1):
                cell = ws.cell(row=row, column=c)
                cell.border = thin_border
                cell.alignment = left
                if alt_fill:
                    cell.fill = alt_fill

        # SHEET 1 – VĂN BẢN ĐI
        if loai in ('all', 'van_ban_di'):
            ws1 = wb.active
            ws1.title = 'Văn bản đi'

            ws1.merge_cells('A1:G1')
            ws1['A1'] = f'BÁO CÁO VĂN BẢN ĐI – THÁNG {thang}/{nam}'
            ws1['A1'].font = title_font
            ws1['A1'].alignment = center

            ws1.merge_cells('A2:G2')
            ws1['A2'] = f'Xuất ngày: {now.strftime("%d/%m/%Y %H:%M")}'
            ws1['A2'].alignment = center

            headers_di = ['STT', 'Số hiệu', 'Tiêu đề', 'Trích yếu', 'Ngày đi', 'Người ký', 'Trạng thái']
            for i, h in enumerate(headers_di, 1):
                ws1.cell(row=4, column=i, value=h)
            style_header_row(ws1, 4, len(headers_di), header_fill_xh)

            domain = [
                ('ngay_di', '>=', f'{nam}-{thang:02d}-01'),
                ('ngay_di', '<=', f'{nam}-{thang:02d}-31'),
            ]
            van_ban_di = request.env['van_ban_di'].search(domain, order='ngay_di asc')

            for idx, vb in enumerate(van_ban_di, 1):
                row = idx + 4
                ws1.cell(row=row, column=1, value=idx)
                ws1.cell(row=row, column=2, value=vb.so_hieu or '')
                ws1.cell(row=row, column=3, value=vb.tieu_de or '')
                ws1.cell(row=row, column=4, value=vb.trich_yeu or '')
                ws1.cell(row=row, column=5, value=str(vb.ngay_di) if vb.ngay_di else '')
                ws1.cell(row=row, column=6, value=vb.id_nguoi_phat_hanh.ho_ten if vb.id_nguoi_phat_hanh else '')
                ws1.cell(row=row, column=7, value=dict(vb._fields['state'].selection).get(vb.state, '') if hasattr(vb, 'state') else '')
                style_data_row(ws1, row, len(headers_di), idx % 2 == 0)

            tong_row = len(van_ban_di) + 5
            ws1.cell(row=tong_row, column=1, value='TỔNG CỘNG').font = Font(bold=True)
            ws1.cell(row=tong_row, column=2, value=len(van_ban_di))

            ws1.column_dimensions['A'].width = 6
            ws1.column_dimensions['B'].width = 14
            ws1.column_dimensions['C'].width = 30
            ws1.column_dimensions['D'].width = 35
            ws1.column_dimensions['E'].width = 12
            ws1.column_dimensions['F'].width = 20
            ws1.column_dimensions['G'].width = 14
            ws1.row_dimensions[1].height = 25

        # SHEET 2 – VĂN BẢN ĐẾN
        if loai in ('all', 'van_ban_den'):
            ws2 = wb.create_sheet('Văn bản đến')
            ws2.merge_cells('A1:G1')
            ws2['A1'] = f'BÁO CÁO VĂN BẢN ĐẾN – THÁNG {thang}/{nam}'
            ws2['A1'].font = title_font
            ws2['A1'].alignment = center

            headers_den = ['STT', 'Số hiệu', 'Tiêu đề', 'Trích yếu', 'Ngày đến', 'Cơ quan ban hành', 'Người nhận']
            for i, h in enumerate(headers_den, 1):
                ws2.cell(row=4, column=i, value=h)
            style_header_row(ws2, 4, len(headers_den), header_fill_xh)

            domain_den = [
                ('ngay_den', '>=', f'{nam}-{thang:02d}-01'),
                ('ngay_den', '<=', f'{nam}-{thang:02d}-31'),
            ]
            van_ban_den = request.env['van_ban_den'].search(domain_den, order='ngay_den asc')

            for idx, vb in enumerate(van_ban_den, 1):
                row = idx + 4
                ws2.cell(row=row, column=1, value=idx)
                ws2.cell(row=row, column=2, value=vb.so_hieu or '')
                ws2.cell(row=row, column=3, value=vb.tieu_de or '')
                ws2.cell(row=row, column=4, value=vb.trich_yeu or '')
                ws2.cell(row=row, column=5, value=str(vb.ngay_den) if vb.ngay_den else '')
                ws2.cell(row=row, column=6, value=vb.co_quan_ban_hanh or '')
                ws2.cell(row=row, column=7, value=vb.id_nguoi_nhan.ho_ten if vb.id_nguoi_nhan else '')
                style_data_row(ws2, row, len(headers_den), idx % 2 == 0)

            ws2.column_dimensions['A'].width = 6
            ws2.column_dimensions['B'].width = 14
            ws2.column_dimensions['C'].width = 30
            ws2.column_dimensions['D'].width = 35
            ws2.column_dimensions['E'].width = 12
            ws2.column_dimensions['F'].width = 22
            ws2.column_dimensions['G'].width = 20

        # SHEET 3 – CÔNG VIỆC
        if loai in ('all', 'cong_viec'):
            ws3 = wb.create_sheet('Công việc')
            ws3.merge_cells('A1:H1')
            ws3['A1'] = f'BÁO CÁO CÔNG VIỆC – THÁNG {thang}/{nam}'
            ws3['A1'].font = title_font
            ws3['A1'].alignment = center

            headers_cv = ['STT', 'Tên công việc', 'Chỉ đạo', 'Chủ trì', 'Ngày tạo', 'Hạn xử lý', 'Ngày hoàn thành', 'Trạng thái']
            for i, h in enumerate(headers_cv, 1):
                ws3.cell(row=4, column=i, value=h)
            style_header_row(ws3, 4, len(headers_cv), header_fill_cv)

            domain_cv = [
                ('ngay_tao', '>=', f'{nam}-{thang:02d}-01'),
                ('ngay_tao', '<=', f'{nam}-{thang:02d}-31'),
            ]
            cong_viec = request.env['cong.viec'].search(domain_cv, order='ngay_tao asc')

            for idx, cv in enumerate(cong_viec, 1):
                row = idx + 4
                ws3.cell(row=row, column=1, value=idx)
                ws3.cell(row=row, column=2, value=cv.ten_cong_viec or '')
                ws3.cell(row=row, column=3, value=cv.chi_dao.ho_ten if cv.chi_dao else '')
                ws3.cell(row=row, column=4, value=cv.chu_tri_giai_quyet.ho_ten if cv.chu_tri_giai_quyet else '')
                ws3.cell(row=row, column=5, value=str(cv.ngay_tao) if cv.ngay_tao else '')
                ws3.cell(row=row, column=6, value=str(cv.han_xu_ly) if cv.han_xu_ly else '')
                ws3.cell(row=row, column=7, value=str(cv.ngay_hoan_thanh) if cv.ngay_hoan_thanh else '')
                ws3.cell(row=row, column=8, value=cv.trang_thai.ten_trang_thai if getattr(cv, 'trang_thai', False) else '')
                style_data_row(ws3, row, len(headers_cv), idx % 2 == 0)

            ws3.column_dimensions['A'].width = 6
            ws3.column_dimensions['B'].width = 30
            ws3.column_dimensions['C'].width = 20
            ws3.column_dimensions['D'].width = 20
            ws3.column_dimensions['E'].width = 12
            ws3.column_dimensions['F'].width = 12
            ws3.column_dimensions['G'].width = 18
            ws3.column_dimensions['H'].width = 18

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f'BaoCao_QLVB_{thang:02d}_{nam}.xlsx'
        return request.make_response(
            output.read(),
            headers=[
                ('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'),
                ('Content-Disposition', f'attachment; filename="{filename}"'),
            ],
        )

